from openpyxl import load_workbook
import json

"""
xlsxds file format:
{
	"xlsxfile":"./data.xlsx",
	"data_from":7,
	"data_sheet":"Sheet1",
	"label_at",1,
	"name_at":null,
	"datatype_at":2,
	"ioattrs_at":3,
	"listhide_at":4,
	"inputhide_at":5,
	"frozen_at":6
}
"""

class XLSXData:
	def __init__(self,path,desc):
		self.desc = desc
		self.xlsxfile = path
		self.workbook = load_workbook(self.xlsxfile)
		self.ws = self.workbook[self.desc['data_sheet']]
	
	def getBaseFieldsInfo(self):
		ws = self.workbook[self.desc['data_sheet']]
		ret = []
		for y in range(1,ws.max_column+1):
			r = {
				'name':self._fieldName(ws,y),
				'label':self._fieldLabel(ws,y),
				'type':self._fieldType(ws,y),
				'listhide':self._isListHide(ws,y),
				'inputhide':self._isInputHide(ws,y),
				'frozen':self._isFrozen(ws,y)
			}
			r.update(self._fieldIOattrs(ws,y))
			ret.append(r)
		return ret

	def _fieldName(self,ws,i):
		x = self.desc.get('name_at')
		if x is not None:
			return ws.cell(x,i).value
		return 'f' + str(i)
	def _fieldLabel(self,ws,i):
		x = self.desc.get('label_at',1)
		if x is not None:
			return ws.cell(x,i).value
		return 'f' + str(i)
	def _fieldType(self,ws,i):
		x = self.desc.get('datatype_at')
		if x is not None:
			return ws.cell(x,i).value
		return 'str'
	def _fieldIOattrs(self,ws,i):
		x = self.desc.get('ioattrs_at')
		if x is not None:
			t = ws.cell(x,i).value
			if t is not None:
				try:
					return json.loads(t,'utf-8')
				except Exception as e:
					print('xlsxData.py:field=',i,'t=',t,'error')
		return {}
	def _isFrozen(self,ws,i):
		x = self.desc.get('frozen_at')
		if x is not None:
			t = ws.cell(x,y).value
			if t == 'Y' or t == 'y':
				return True
		return False
	def _isListHide(self,ws,i):
		x = self.desc.get('listhide_at')
		if x is not None:
			t = ws.cell(x,i).value
			if t == 'Y' or t == 'y':
				return True
		return False
	def _isInputHide(self,ws,i):
		x = self.desc.get('inputhide_at')
		if x is not None:
			t = ws.cell(x,i).value
			if t == 'Y' or t == 'y':
				return True
		return False

	def getPeriodData(self,min_r,max_r):
		ws = self.ws
		rows = []
		assert(min_r >= self.desc.get('data_from',2))
		if max_r > ws.max_row:
			max_r = ws.max_row + 1;
	
		if min_r <= max_r:
			x = min_r;
			while x < max_r:
				d = {}
				for y in range(1,ws.max_column+1):
					name = self._fieldName(ws,y)
					d.update({name:ws.cell(column=y,row=x).value})
				rows.append(d)
				x = x + 1
		return rows
		
	def getArgumentsDesc(self,ns,request):
		return None

	def getData(self,ns):
		ws = self.ws
		min_r = self.desc.get('data_from',2)
		return self.getPeriodData(min_r,ws.max_row + 1)

	def getPagingData(self,ns):
		rows = int(ns.get('rows',50))
		page = int(ns.get('page',1))
		d1 = self.desc.get('data_from',2)
		min_r = (page - 1) * rows + d1
		max_r = page * rows + d1 + 1
		rows = self.getPeriodData(min_r,max_r)
		ret = {
			'total':self.ws.max_row - d1,
			'rows':rows
		}
		return ret
		
